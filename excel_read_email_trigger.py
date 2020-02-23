import openpyxl
import datetime
from email_sending import *

# Create Sheet Object & get the sheet name
wb = openpyxl.load_workbook('payment.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# Get Current date
now = datetime.datetime.now()
current_date = now.strftime("%m/%d/%y")

# Identify Sender's credentials and Email.
senders_mail = str(sheet.cell(row=1, column=1).value)
credential = str(sheet.cell(row=2, column=1).value)

#get_num_of_rows = input('Enter number of rows in the sheet')

# for i in range(2, get_num_of_rows + 1):

# Loop through all data of excel sheet and get the required details
for i in range(2, 10):
    if (sheet.cell(row=i, column=5).value) != None:
        if datetime.datetime.strftime(
                sheet.cell(row=i, column=6).value, "%m/%d/%y") <= current_date:
            name = 'Hey,' + str(sheet.cell(row=i, column=3).value) + '! '
            bal_due = '\n You have a balance due of $' + str(sheet.cell(row=i, column=5).value)
            msg = name + ' ' + bal_due
            recipient_mail = sheet.cell(row=i, column=4).value
            #print(sheet.cell(row=i, column=5).value)
            send_email(senders_mail, recipient_mail, credential, msg)
